#!/usr/bin/env python3
"""
Generate grading report CSV for Canvas upload.
Matches evaluation results with student information from submissions CSV.
"""

import json
import csv
from pathlib import Path

def load_submissions(csv_path):
    """Load student submissions from CSV."""
    students = []
    with open(csv_path, 'r') as f:
        reader = csv.DictReader(f)
        for row in reader:
            url = row['submission_url']
            # Extract repo name from GitHub URL
            if 'github.com' in url:
                # Handle URLs like: https://github.com/FAU-AI-HootCamp-Summer-2026/week3-jaltamirano6933
                # or: https://github.com/abarahonacab2025-code/week3-abarahonacab2025-code.git
                parts = url.split('/')
                repo_name = parts[-1].replace('.git', '')
                full_path = '/'.join(parts[-2:])
                students.append({
                    'name': row['student_name'],
                    'email': row['email'],
                    'url': url,
                    'repo_name': repo_name,
                    'full_path': full_path,
                    'is_direct_url': False
                })
            else:
                # Handle non-GitHub URLs (Vercel, etc.)
                students.append({
                    'name': row['student_name'],
                    'email': row['email'],
                    'url': url,
                    'repo_name': None,
                    'full_path': None,
                    'is_direct_url': True
                })
    return students

def load_evaluation_results(json_path):
    """Load evaluation results from JSON."""
    with open(json_path, 'r') as f:
        return json.load(f)

def extract_repo_name(repo_path):
    """Extract repo name from repo path, removing organization prefix."""
    full_name = repo_path.split('/')[-1]
    # Remove organization prefix if present (e.g., FAU-AI-HootCamp-Summer-2026_week3-xxx -> week3-xxx)
    if '_' in full_name:
        parts = full_name.split('_')
        # Find the part that starts with 'week' (week2, week3, etc.)
        for i, part in enumerate(parts):
            if part.startswith('week'):
                return '_'.join(parts[i:])
    return full_name

def clean_github_url(url):
    """Clean up GitHub URL by removing .git and /tree/main suffixes."""
    if not url or not isinstance(url, str):
        return url
    
    # Remove .git suffix
    url = url.replace('.git', '')
    
    # Remove /tree/main suffix
    url = url.replace('/tree/main', '')
    
    # Remove trailing slash
    url = url.rstrip('/')
    
    return url

def extract_urls_from_explanation(explanation):
    """Pull http(s) URLs out of gate explanation text/dicts/lists."""
    import re
    url_pattern = r'https?://[^\s\)\]\}"\'<>]+'
    urls = []

    def collect(value):
        if isinstance(value, str):
            urls.extend(re.findall(url_pattern, value))
        elif isinstance(value, dict):
            for v in value.values():
                collect(v)
        elif isinstance(value, list):
            for v in value:
                collect(v)

    collect(explanation)
    # Strip trailing punctuation left by OCR/LLM text
    cleaned = []
    for url in urls:
        cleaned.append(url.rstrip('.,);:\'"`'))
    return cleaned

LIVE_HOST_HINTS = (
    'vercel.app', 'netlify.app', 'onrender.com', 'railway.app',
    'fly.dev', 'pages.dev', 'web.app', 'firebaseapp.com', 'hosted.app',
    'herokuapp.com', 'amplifyapp.com', 'azurewebsites', 'cloudfront',
    'github.io',
)

LIVE_SKIP_HINTS = (
    'github.com/', 'youtu', 'npmjs', 'supabase.co', 'openai.com',
    'googleapis', 'localhost', '127.0.0.1', 'example.com', 'docs.',
    'console.', 'portal.azure', 'parseapi.back4app', 'back4app.com',
    'netlify.com', 'vercel.com', 'app.netlify.com', 'YOUR-SITE',
    'your-database', 'your-site', 'your-project', 'integrate.api.nvidia',
    'alpaca.markets', 'nextjs.org', 'oxc.rs', 'git-scm.com',
    'aistudio.google', 'fauengtrussed', 'apod.nasa.gov', 'drive.google.com',
    'loom.com', 'mongodb.com', 'anthropic.com', 'platform.openai',
    'img.shields.io', 'user-attachments', 'actions/workflows',
)


def is_live_app_url(url):
    """True for deployed app URLs, false for docs/video/placeholder links."""
    if not url or not isinstance(url, str):
        return False
    lower = url.lower().rstrip('/')
    if any(skip in lower for skip in LIVE_SKIP_HINTS):
        return False
    if any(host in lower for host in LIVE_HOST_HINTS):
        return True

    # Custom domains: accept app.* hosts and known student domains
    host = lower.split('//', 1)[-1].split('/', 1)[0]
    if host.startswith(('app.', 'www.')) and host.count('.') >= 2:
        return True
    if host.endswith(('.com', '.app', '.io', '.dev', '.ai')) and host.count('.') >= 1:
        # Avoid bare vendor marketing sites with no product path unless subdomain
        if host.count('.') >= 2:  # subdomain.domain.tld
            return True
        path = lower.split(host, 1)[-1]
        if path and path not in ('', '/'):
            return True
    return False


def pick_best_live_url(urls):
    """Prefer production app hosts over dashboards/placeholders."""
    candidates = []
    for url in urls:
        cleaned = url.rstrip('/')
        if is_live_app_url(cleaned):
            candidates.append(cleaned)

    if not candidates:
        return ''

    def score(u):
        lower = u.lower()
        # Higher is better
        points = 0
        if any(h in lower for h in (
            'vercel.app', 'netlify.app', 'onrender.com', 'railway.app',
            'fly.dev', 'hosted.app', 'web.app', 'firebaseapp.com'
        )):
            points += 100
        if lower.startswith(('https://app.', 'http://app.')):
            points += 80
        if 'blog.' in lower or '/blog' in lower:
            points -= 50
        if '/api' in lower or '/health' in lower:
            points -= 20
        if '/login' in lower:
            points -= 5
        # Prefer shorter root URLs
        points -= u.count('/') * 2
        points -= len(u) // 20
        return points

    candidates.sort(key=score, reverse=True)
    return candidates[0]


def collect_readme_urls(repo_path):
    """Collect URLs from local README files."""
    if not repo_path:
        return []
    root = Path(repo_path)
    if not root.exists():
        return []

    readme_files = []
    for pattern in ('README*', 'readme*', '**/README.md'):
        readme_files.extend(root.glob(pattern))

    urls = []
    seen = set()
    for readme in readme_files[:8]:
        if not readme.is_file():
            continue
        try:
            text = readme.read_text(errors='ignore')
        except Exception:
            continue
        for url in extract_urls_from_explanation(text):
            if url not in seen:
                seen.add(url)
                urls.append(url)
    return urls


def extract_live_url_from_readme(repo_path):
    """Scan local repo README(s) for a deployed app URL."""
    return pick_best_live_url(collect_readme_urls(repo_path))


def extract_live_url(gate_evaluations, student_url='', repo_path=''):
    """Prefer explicit live URL evidence, then README, then non-GitHub submission URL."""
    # 1) deployment_live explanation
    deployment_eval = gate_evaluations.get('deployment_live', {})
    urls = extract_urls_from_explanation(deployment_eval.get('explanation', ''))
    live = pick_best_live_url(urls)
    if live:
        return live

    # 2) other gate explanations often include the live link (especially README)
    for gate_name in ('readme_completeness', 'demo_video_present', 'ai_integration',
                      'backend_database', 'authentication'):
        gate = gate_evaluations.get(gate_name, {})
        urls = extract_urls_from_explanation(gate.get('explanation', ''))
        live = pick_best_live_url(urls)
        if live:
            return live

    # 3) local README scan
    live = extract_live_url_from_readme(repo_path)
    if live:
        return live

    # 4) non-GitHub submission URL (Vercel/Netlify link submitted directly)
    if student_url and 'github.com' not in student_url.lower() and is_live_app_url(student_url):
        return student_url.rstrip('/')
    return ''


VIDEO_HOST_HINTS = (
    'youtu', 'vimeo', 'loom', 'drive.google', 'dropbox',
    'streamable', 'wistia', 'panopto', 'mediasite',
)


def is_video_url(url):
    if not url or not isinstance(url, str):
        return False
    lower = url.lower()
    if any(host in lower for host in VIDEO_HOST_HINTS):
        return True
    # Local/repo-hosted demo videos
    if lower.endswith(('.mp4', '.mov', '.webm', '.m4v')):
        return True
    return False


def pick_best_video_url(urls):
    for url in urls:
        cleaned = url.rstrip('.,);:\'"`')
        if is_video_url(cleaned):
            return cleaned
    return ''


def extract_video_url(gate_evaluations, repo_path=''):
    """Extract demo video URL from gate text, then README."""
    # 1) demo_video_present explanation
    demo_video_eval = gate_evaluations.get('demo_video_present', {})
    urls = extract_urls_from_explanation(demo_video_eval.get('explanation', ''))
    video = pick_best_video_url(urls)
    if video:
        return video

    # 2) other gate explanations (README completeness often has the link)
    for gate_name in ('readme_completeness', 'deployment_live'):
        gate = gate_evaluations.get(gate_name, {})
        urls = extract_urls_from_explanation(gate.get('explanation', ''))
        video = pick_best_video_url(urls)
        if video:
            return video

    # 3) local README scan — even if LLM said present without quoting the URL
    return pick_best_video_url(collect_readme_urls(repo_path))

def generate_grading_report(evaluation_results, students):
    """Generate grading report matching repos to students."""
    report = []
    matched_students = set()
    
    # Create a mapping from repo names to students
    repo_to_student = {}
    for student in students:
        if student['repo_name']:
            repo_to_student[student['repo_name']] = student
        if student['full_path']:
            repo_to_student[student['full_path']] = student
    
    # Manual mappings for special cases
    manual_mappings = {
        'leonfallett10-netizen_studyflow': {'name': 'Leon Fallett', 'email': 'lfallett2021@fau.edu', 'url': 'https://github.com/leonfallett10-netizen/studyflow'},
        'week3-olmard-guillaume': {'name': 'Olmard Guillaume', 'email': 'oguillaume2025@fau.edu', 'url': 'https://github.com/FAU-AI-HootCamp-Summer-2026/week3-olmard-guillaume'},
        'week3-CruzIsaiah': {'name': 'Isaiah Cruz', 'email': 'icruz2020@fau.edu', 'url': 'https://github.com/FAU-AI-HootCamp-Summer-2026/week3-CruzIsaiah'},
        'week3-gissellapam': {'name': 'Gissella Palomino', 'email': 'grodri32@fau.edu', 'url': 'https://github.com/FAU-AI-HootCamp-Summer-2026/week3-gissellapam'},
        'week2-Anthonyhage99': {'name': 'Anthony Hage', 'email': 'ahage2024@fau.edu', 'url': 'https://github.com/FAU-AI-HootCamp-Summer-2026/week2-Anthonyhage99.git'},
        'week3-Yaswanth-Kotipalli': {'name': 'Yaswanth Durga Kiran Kotipalli', 'email': 'ykotipalli2024@fau.edu', 'url': 'https://github.com/FAU-AI-HootCamp-Summer-2026/week3-Yaswanth-Kotipalli/tree/main'}
    }
    
    for result in evaluation_results:
        repo_name = extract_repo_name(result['repo_path'])
        student = None
        
        # Check manual mappings first
        if repo_name in manual_mappings:
            student = manual_mappings[repo_name]
        # Try exact match
        elif repo_name in repo_to_student:
            student = repo_to_student[repo_name]
        else:
            # Try matching with organization prefix
            for key, value in repo_to_student.items():
                if repo_name == key.split('/')[-1]:
                    student = value
                    break
                if repo_name.lower() in key.lower():
                    student = value
                    break
        
        if student:
            # Skip if this student was already matched (avoid duplicates)
            if student['name'] in matched_students:
                continue
            matched_students.add(student['name'])
            
            gate_evaluations = result.get('gate_evaluations', {})
            if not gate_evaluations:
                status = 'FAIL'
                reason = 'Incomplete evaluation (missing gate data) — re-run hc_evaluate'
            elif result['gate_pass']:
                status = 'PASS'
                reason = 'All gate requirements met'
            else:
                status = 'FAIL'
                missing = result['gate_decision'].get('missing_requirements', [])
                ast_issues = result['gate_decision'].get('ast_issues', [])
                reasons = []
                if missing:
                    reasons.append(f"Missing: {', '.join(missing)}")
                # Only include AST issues that correspond to failed gates
                if ast_issues:
                    # Map AST issues to their corresponding gates
                    ast_issue_to_gate = {
                        'No database library (Supabase, Back4App, MongoDB, or SQL/ORM) detected in AST': 'backend_database',
                        'No auth library detected in AST': 'authentication',
                        'No AI library detected in AST': 'ai_integration',
                    }
                    # Only include AST issues for gates that are actually missing
                    relevant_ast_issues = []
                    for issue in ast_issues:
                        gate_name = ast_issue_to_gate.get(issue)
                        if gate_name and gate_name in missing:
                            relevant_ast_issues.append(issue.replace(' in AST', ''))
                    if relevant_ast_issues:
                        reasons.append(', '.join(relevant_ast_issues))
                reason = '; '.join(reasons) if reasons else 'Failed gate evaluation'
            
            # Extract individual gate statuses
            gates = {
                'AI Integration': gate_evaluations.get('ai_integration', {}).get('present', False),
                'Backend Database': gate_evaluations.get('backend_database', {}).get('present', False),
                'Authentication': gate_evaluations.get('authentication', {}).get('present', False),
                'Readme Completeness': gate_evaluations.get('readme_completeness', {}).get('present', False),
                'Deployment Live': gate_evaluations.get('deployment_live', {}).get('present', False),
                'Demo Video Present': gate_evaluations.get('demo_video_present', {}).get('present', False)
            }
            
            video_url = extract_video_url(gate_evaluations, result.get('repo_path', ''))
            live_url = extract_live_url(
                gate_evaluations,
                student.get('url', ''),
                result.get('repo_path', ''),
            )
            
            report.append({
                'Student Name': student['name'],
                'Email': student['email'],
                'Status': status,
                'Reason': reason,
                'GitHub URL': clean_github_url(student['url']),
                'Live URL': live_url,
                'Video URL': video_url,
                **{k: 'PASS' if v else 'FAIL' for k, v in gates.items()}
            })
        else:
            # No matching student found
            report.append({
                'Student Name': 'UNKNOWN',
                'Email': 'unknown@fau.edu',
                'Status': 'FAIL' if not result['gate_pass'] else 'PASS',
                'Reason': 'No matching student in submissions CSV',
                'GitHub URL': clean_github_url(result['repo_path']),
                'Live URL': '',
                'Video URL': '',
                'AI Integration': 'N/A',
                'Backend Database': 'N/A',
                'Authentication': 'N/A',
                'Readme Completeness': 'N/A',
                'Deployment Live': 'N/A',
                'Demo Video Present': 'N/A'
            })
    
    return report

def short_url_label(url, kind='live'):
    """Compact display text for URL columns (full URL stays in hyperlink)."""
    if not url or not isinstance(url, str):
        return ''
    cleaned = url.rstrip('/')
    lower = cleaned.lower()
    if kind == 'github':
        return cleaned.split('/')[-1] or cleaned
    if kind == 'video':
        if 'youtu.be/' in lower:
            return 'YouTube'
        if 'youtube.com' in lower:
            return 'YouTube'
        if 'loom.com' in lower:
            return 'Loom'
        if 'drive.google' in lower:
            return 'Google Drive'
        if 'vimeo.com' in lower:
            return 'Vimeo'
        if 'dropbox.com' in lower:
            return 'Dropbox'
        return 'Video'
    # live
    host = cleaned.split('//', 1)[-1].split('/', 1)[0]
    if host.startswith('www.'):
        host = host[4:]
    return host


def save_grading_report(report, output_path):
    """Save grading report to XLSX with color coding and clickable URLs."""
    try:
        import pandas as pd
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("Error: pandas or openpyxl not installed. Install with: pip install pandas openpyxl")
        return
    
    # Convert to DataFrame
    fieldnames = [
        'Student Name', 'Email', 'Status', 'Reason',
        'AI Integration', 'Backend Database', 'Authentication',
        'Readme Completeness', 'Deployment Live', 'Demo Video Present',
        'GitHub URL', 'Live URL', 'Video URL'
    ]
    df = pd.DataFrame(report, columns=fieldnames)
    
    # Sort by student name
    df = df.sort_values(by='Student Name')
    
    # Save to XLSX
    df.to_excel(output_path, index=False, engine='openpyxl')
    
    # Apply color coding and hyperlinks
    wb = load_workbook(output_path)
    ws = wb.active
    
    # Define colors
    pass_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")  # Light green
    fail_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")  # Stronger red
    
    # Columns to color code (Status and all gate columns)
    status_col_idx = fieldnames.index('Status') + 1  # 1-based index
    gate_columns = ['AI Integration', 'Backend Database', 'Authentication',
                   'Readme Completeness', 'Deployment Live', 'Demo Video Present']
    gate_col_indices = [fieldnames.index(col) + 1 for col in gate_columns]
    
    # URL columns
    github_url_col_idx = fieldnames.index('GitHub URL') + 1
    live_url_col_idx = fieldnames.index('Live URL') + 1
    video_url_col_idx = fieldnames.index('Video URL') + 1
    
    # Adjust column widths (narrower)
    column_widths = {
        'Student Name': 20,
        'Email': 25,
        'Status': 10,
        'Reason': 40,
        'AI Integration': 10,
        'Backend Database': 12,
        'Authentication': 10,
        'Readme Completeness': 12,
        'Deployment Live': 10,
        'Demo Video Present': 12,
        'GitHub URL': 25,
        'Live URL': 40,
        'Video URL': 25
    }
    
    for col_name, width in column_widths.items():
        col_idx = fieldnames.index(col_name) + 1
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    # Make header row taller for text wrapping and enable wrap text with center alignment
    ws.row_dimensions[1].height = 50
    from openpyxl.styles import Alignment, Font
    link_font = Font(color="0563C1", underline="single")
    for col_idx in range(1, len(fieldnames) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
    
    # Apply color coding and hyperlinks to each row (skip header)
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        # Status column
        status_cell = row[status_col_idx - 1]
        if status_cell.value == 'PASS':
            status_cell.fill = pass_fill
        elif status_cell.value == 'FAIL':
            status_cell.fill = fail_fill
        status_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Gate columns
        for col_idx in gate_col_indices:
            cell = row[col_idx - 1]
            if cell.value == 'PASS':
                cell.fill = pass_fill
            elif cell.value == 'FAIL':
                cell.fill = fail_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Make GitHub URL clickable with repo name as display text
        github_cell = row[github_url_col_idx - 1]
        if github_cell.value and isinstance(github_cell.value, str) and github_cell.value.startswith('http'):
            url = github_cell.value
            github_cell.value = short_url_label(url, 'github')
            github_cell.hyperlink = url
            github_cell.font = link_font
        
        # Make Live URL clickable with host as display text
        live_cell = row[live_url_col_idx - 1]
        if live_cell.value and isinstance(live_cell.value, str) and live_cell.value.startswith('http'):
            url = live_cell.value
            live_cell.value = short_url_label(url, 'live')
            live_cell.hyperlink = url
            live_cell.font = link_font
        
        # Make Video URL clickable with short label
        video_cell = row[video_url_col_idx - 1]
        if video_cell.value and isinstance(video_cell.value, str) and video_cell.value.startswith('http'):
            url = video_cell.value
            video_cell.value = short_url_label(url, 'video')
            video_cell.hyperlink = url
            video_cell.font = link_font
    
    # Save with formatting
    wb.save(output_path)
    
    print(f"Grading report saved to {output_path}")
    print(f"Total entries: {len(report)}")
    
    # Print summary
    passed = sum(1 for r in report if r['Status'] == 'PASS')
    failed = sum(1 for r in report if r['Status'] == 'FAIL')
    print(f"Passed: {passed}")
    print(f"Failed: {failed}")

def main():
    # Paths
    submissions_csv = 'submissions/week3_submissions_updated.csv'
    evaluation_json = 'results/evaluation_results.json'
    output_file = 'results/hootcamp_grading_report.xlsx'
    
    # Load data
    print("Loading submissions...")
    students = load_submissions(submissions_csv)
    print(f"Found {len(students)} students")
    
    print("Loading evaluation results...")
    results = load_evaluation_results(evaluation_json)
    print(f"Found {len(results)} evaluation results")
    
    # Generate report
    print("Generating grading report...")
    report = generate_grading_report(results, students)
    
    # Save report
    save_grading_report(report, output_file)

if __name__ == '__main__':
    main()
