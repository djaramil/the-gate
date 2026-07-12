#!/usr/bin/env python3
"""
Extract project information from FAU-AI-Hootcamp-Problem-Set.pdf and create Excel file.
"""

import re
import sys
from pathlib import Path

try:
    import PyPDF2
except ImportError:
    print("Installing PyPDF2...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "PyPDF2"])
    import PyPDF2

try:
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Font
except ImportError:
    print("Installing pandas and openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "pandas openpyxl"])
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Font

def extract_text_from_pdf(pdf_path):
    """Extract all text from PDF file."""
    text = ""
    with open(pdf_path, 'rb') as file:
        reader = PyPDF2.PdfReader(file)
        for page in reader.pages:
            text += page.extract_text() + "\n"
    return text

def parse_projects(text):
    """Parse project information from extracted text."""
    projects = []
    
    # Split text into lines and process
    lines = text.split('\n')
    
    current_project = None
    current_category = None
    
    for line in lines:
        line = line.strip()
        
        # Look for category headers (based on the PDF description)
        if re.match(r'^(Lead Problems|Live or Evergreen|Ready-to-Use|Industry-Sponsored)', line, re.IGNORECASE):
            current_category = line
            continue
            
        # Look for project entries - they start with page numbers
        # Pattern: "2  ACL Caregiver AI Prize Challenge Category Lead problem (live)" or similar
        page_match = re.match(r'^(\d+)\s+(.+)$', line)
        if page_match:
            # Save previous project if exists
            if current_project and current_project.get('title'):
                projects.append(current_project)
            
            # Start new project - extract only the title before "Category"
            full_text = page_match.group(2)
            # Split at "Category" to get just the title
            if 'Category' in full_text:
                title = full_text.split('Category')[0].strip()
            else:
                title = full_text.strip()
            
            # Clean up title by removing text after keywords that indicate start of descriptions
            title_end_keywords = [
                'Mentorship and Support', 'Notes', 'Data and Resources', 'Sponsor:',
                'Resources and Links', 'Intellectual property', 'Showcase eligibility',
                'System constraints', 'Deployment and execution', 'Requirements-to-Task',
                'Suggested epics', 'User stories', 'Engineering tasks', 'Development milestones',
                'Recommended implementation', 'To support development', 'These materials are intended',
                'What are the deployment steps', 'The system generates', 'Success will be measured'
            ]
            for keyword in title_end_keywords:
                if keyword in title:
                    title = title.split(keyword)[0].strip()
            
            # Skip non-project entries (like the PDF header page and continuation text)
            skip_keywords = ['Project Menu', 'Hootcamp is a', 'summer program', 'participant works',
                            'Resources and Links', 'Mentship and Support', 'Notes', 'Intellectual property',
                            'Showcase eligibility', 'System constraints', 'Security and API', 'Build instructions',
                            'Deployment and execution', 'Requirements-to-Task', 'Suggested epics', 'User stories',
                            'Engineering tasks', 'Development milestones', 'Recommended implementation',
                            'Data and Resources', 'To support development', 'These materials are intended',
                            'Sponsor:', 'Principal Architect', 'President', 'CEO']
            
            # Also skip entries that start with bullet points or are very long (likely continuation text)
            if title.startswith('•') or len(title) > 200:
                current_project = None
                continue
            
            # Skip empty or invalid titles
            if not title or len(title) < 5 or title in ['"', '"', '"']:
                current_project = None
                continue
                
            if any(keyword.lower() in title.lower() for keyword in skip_keywords):
                current_project = None
                continue
            
            current_project = {
                'page': page_match.group(1),
                'category': current_category or 'General',
                'title': title,
                'Student Name': '',
                'Student Email': '',
                'Student Name 2': '',
                'Student Email 2': ''
            }
        elif current_project:
            # Parse project details
            if line.startswith('Category'):
                current_project['category'] = line.replace('Category', '').strip()
    
    # Don't forget the last project
    if current_project and current_project.get('title'):
        projects.append(current_project)
    
    return projects

def main():
    pdf_path = '/Users/yoda26/Documents/FAU/HootCamp-Summer2026/the-gate/FAU-AI-Hootcamp-Problem-Set.pdf'
    output_path = '/Users/yoda26/Documents/FAU/HootCamp-Summer2026/the-gate/AI-Hootcamp-Projects.xlsx'
    
    print(f"Extracting text from {pdf_path}...")
    text = extract_text_from_pdf(pdf_path)
    
    print("Parsing project information...")
    projects = parse_projects(text)
    
    print(f"Found {len(projects)} projects")
    
    # Create DataFrame
    if projects:
        df = pd.DataFrame(projects)
        
        # Define the columns we want (with capital letters for display)
        expected_columns = ['Page', 'Title', 'Student Name', 'Student Email', 'Student Name 2', 'Student Email 2']
        
        # Rename existing columns to match expected format
        df = df.rename(columns={'page': 'Page', 'title': 'Title'})
        
        # Add missing columns with empty strings
        for col in expected_columns:
            if col not in df.columns:
                df[col] = ''
        
        # Select only the columns we want
        df = df[expected_columns]
        
        df.to_excel(output_path, index=False, engine='openpyxl')
        
        # Apply formatting to the Excel file
        wb = load_workbook(output_path)
        ws = wb.active
        
        # Make header row bold and larger font
        header_font = Font(bold=True, size=14)
        for cell in ws[1]:  # First row (headers)
            cell.font = header_font

        # Make all cells 14 point font
        body_font = Font(size=14)
        for row in ws.iter_rows(min_row=2):  # All rows except header
            for cell in row:
                cell.font = body_font

        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Set specific widths for Student Name and Student Email columns
        ws.column_dimensions['C'].width = 25  # Student Name
        ws.column_dimensions['D'].width = 30  # Student Email
        ws.column_dimensions['E'].width = 25  # Student Name 2
        ws.column_dimensions['F'].width = 30  # Student Email 2
        
        wb.save(output_path)
        print(f"Excel file saved to {output_path}")
        
        # Print summary
        print("\nProjects found:")
        for i, project in enumerate(projects, 1):
            print(f"{i}. {project['title']}")
    else:
        print("No projects found. The PDF structure might be different than expected.")
        print("\nFirst 2000 characters of extracted text:")
        print(text[:2000])
        print("\n\nTotal text length:", len(text))

if __name__ == '__main__':
    main()
